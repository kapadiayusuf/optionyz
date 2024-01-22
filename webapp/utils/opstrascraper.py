
import csv
from xmlrpc.client import DateTime
import openpyxl
import pandas as pd
from datetime import date,datetime , timedelta
import time
import json
import requests
import random
import os
import difflib

from bs4 import BeautifulSoup

from selenium import webdriver
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.edge.options import Options as EdgeOptions

''' from webapp.utils.apimanager import ApiManager


apimgr = ApiManager()
 '''
    
class OpstraScraper():
    
    expiry_dates = [ '25JAN2024', '29FEB2024', '28MAR2024', '25APR2024', '30MAY2024', '27JUN2024', '25JUL2024', 
                    '29AUG2024', '26SEP2024', '31OCT2024', '28NOV2024', '26DEC2024']
    
    
    def getFilePath(self):
        # if("E:" in os.getcwd()):
        #     #return "/Users/noaman/Documents/DATA/dev/code/Django/optionstrader/webapp/utils/"
        #     #return "/Users/noam/Documents/dev/code/django/optionstrader/webapp/utils/"
        #     return "/optionstrader/webapp/utils/"
        # else:
        #     return "/var/www/optionstrader/webapp/utils/"
        return "C://django_projects/code/optionstrader/webapp/utils/"    
   
    
    
    def write_to_CSV(self, filename, data):
        
        with open(  filename, 'w', newline='') as csvfile:
            fieldnames = data[0].keys()
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames , lineterminator='\n')

            writer.writeheader()
            writer.writerows(data)

    def getDriver(self) :        #wdPath = 'webdriver/msedgedriver.exe'):
        webdriver_path =  'webdriver/msedgedriver.exe'


        # Create Edge WebDriver with headless option
        #edge_service = EdgeService(webdriver_path)
        edge_options = EdgeOptions()
        edge_options.use_chromium = True  # Use Chromium-based Edge
        edge_options.headless = True



        # Initialize the Selenium WebDriver 
        driver =  webdriver.Chrome(executable_path= webdriver_path , options=edge_options)  #    .Chrome(executable_path='E:/chrome-win32/chrome.exe')
        #webdriver.Chrome(options= edge_options)
        #webdriver.Chrome(executable_path= webdriver_path , options=edge_options)  #    .Chrome(executable_path='E:/chrome-win32/chrome.exe')
        #driver = webdriver.Firefox(executable_path='C:/Program Files/Mozilla Firefox/firefox.exe')

            
        return driver      

    def  getDataFromLink(self,driver, link, dict_name) : 
        # Navigate to the desired URL
        driver.get(link)  # json data

        # Wait for some time to allow the page to load 
        time.sleep(5)

        page_source = driver.page_source
        # Parse the HTML string with BeautifulSoup
        soup = BeautifulSoup(page_source, 'html.parser')

        # Find the <div> element with id "json-data"
        json_body = soup.find('body')
        json_div = json_body.find('div', {'hidden': 'true'})


        # Check if the <div> element exists and has content
        if json_div and json_div.string:
            # Extract the JSON data
            json_data = json_div.string

            # Parse the JSON data into a Python dictionary
            json_object = json.loads(json_data)[dict_name]

            # Now, 'json_object' contains the extracted JSON data
            #print(json_object)
        else:
            print("JSON data not found.")


        return json_object


    def login(self):
        driver = self.getDriver()
        # Open the website
        #print(driver.current_url)

        ####################      login              ###############################################
        driver.get("https://sso.definedge.com/auth/realms/definedge/protocol/openid-connect/auth?response_type=code&client_id=opstra&redirect_uri=https://opstra.definedge.com/ssologin&state=e2cf559f-356c-425a-87e3-032097f643d0&login=true&scope=openid")

        time.sleep(5)

        # Find the username and password fields
        username_field = driver.find_element_by_id("username")
        password_field = driver.find_element_by_id("password")

        #  fill the username and password fields
        username_field.send_keys("kapadiayusuf@gmail.com")
        password_field.send_keys("Zain!786")

        # Find and click the login button
        login_button = driver.find_element_by_id("kc-login")
        login_button.click()

        # Wait for some time to allow the login process to complete (adjust the time as needed)
        time.sleep(2)

        return driver


    def writeOptionsData(self):
        link = "https://opstra.definedge.com/api/optionsdashboard/free"
        
        driver = self.login()
        
        json_data = self.getDataFromLink(driver, link, 'optiondata')
        
        driver.close()
        
        # save it to a JSON file if needed
        with open("options.json", "w") as json_file:
            json.dump(json_data, json_file, indent=4)

        csv_file_path = self.getFilePath()+"Output/opstra_options.csv"

        #self.write_to_CSV(csv_file_path, json_data) 
        exl_file_path = self.getFilePath()+"Output/opstra_options.xlsx"
        pd.DataFrame(json_data ).to_excel(exl_file_path)
        
          
        
        
    def  getOpstraData(self, driver, link, dict_name) : 
        # Navigate to the desired URL
        driver.get(link)  # json data

        # Wait for some time to allow the page to load 
        time.sleep(random.randint(5,15))

        page_source = driver.page_source
        # Parse the HTML string with BeautifulSoup
        soup = BeautifulSoup(page_source, 'html.parser')

        # Find the <div> element with id "json-data"
        json_body = soup.find('body')
        json_div = json_body.find('div', {'hidden': 'true'})


        # Check if the <div> element exists and has content
        if json_div and json_div.string:
            # Extract the JSON data
            json_data = json_div.string

            # Parse the JSON data into a Python dictionary
            json_object = json.loads(json_data)[dict_name]

            # Now, 'json_object' contains the extracted JSON data
            #print(json_object)
            return json_object
        else:
            print("JSON data not found.")
            return  []
    
    
    def get_symbol_data(self,driver , symbol = "COALINDIA"):
        #   link  = https://opstra.definedge.com/api/openinterest/futuresopeninterest/COALINDIA&Combined%20OpenInterest
        #   COALINDIA  shouldbe a variable 

        link = "https://opstra.definedge.com/api/openinterest/futuresopeninterest/" + symbol + "&Combined%20OpenInterest"

        data_list = self.getOpstraData(driver, link, 'data')


        ''' # save it to a JSON file if needed
        json_filepath = symbol + "_futuresOI.json"
        with open( json_filepath , "w") as json_file:
            json.dump(data_list, json_file, indent=4)
            '''
            
            
            
        #   Convert the list to a dictionary
        # Define your headers as a list
        headers = ['Timestamp',	'Open',	'High'	,'Low',	'Close',  'OI', 'Futures_Vol',	'BuildUp',	'8', '9',  'Cash Delivery', 'CashDelofVolume']


        last_21_rows = data_list[-21:]

        # Create a dictionary using zip
        data_dict = [dict(zip(headers, row)) for row in last_21_rows]

        

        return data_dict

    def writeFuturesData(self):
        link = "https://opstra.definedge.com/api/openinterest/futuresbuildup"
        
        driver = self.login()
        data_dict = self.getDataFromLink(driver, link, 'data')
        
        driver.close()

        # save it to a JSON file if needed
        with open("futures.json", "w") as json_file:
            json.dump(data_dict, json_file, indent=4)

        csv_file_path = self.getFilePath()+"Output/opstra_futures.csv"

        #self.write_to_CSV( csv_file_path, data_dict) 
        exl_file_path = self.getFilePath()+"Output/opstra_futures.xlsx"
        pd.DataFrame(data_dict ,columns= data_dict[0].keys()).to_excel(exl_file_path)
        
        
    
    def getDataFrame(self ,driver,  symbol = 'COALINDIA'):
        data_dict = self.get_symbol_data(driver, symbol)

        df = pd.DataFrame(data_dict)
        df['SpotChange'] = df['Close'] - df['Close'].shift(1)
        df['Vol_Change'] = df['Futures_Vol'] - df['Futures_Vol'].shift(1)
        df['Vol_Change%'] = ( df['Vol_Change']/df['Futures_Vol'].shift(1) )* 100
        df['OI_Change'] = df['OI'] - df['OI'].shift(1)
        #df['LongBuildup'] = df['BuildUp'].apply(lambda x: df['OI'] if (x =='green' or x=='lightgreen') else 0 )
    
        condition_G_VOL = df['BuildUp'].str.startswith('green' ) & (df['Vol_Change%'] > 0)
        df.loc[condition_G_VOL, 'LongBuildup'] = df.loc[condition_G_VOL , 'OI']
        #df.loc[df['BuildUp'] =='green' , 'LongBuildup'] = df.loc[df['BuildUp'] =='green' , 'OI']
        #df.loc[df['BuildUp'] =='lightgreen' , 'LongBuildup'] =  df.loc[df['BuildUp'] =='lightgreen' , 'OI']

        condition_R_VOL = df['BuildUp'].str.contains('red' , case=False)  &  (df['Vol_Change%'] > 0)
        df.loc[condition_R_VOL , 'ShortBuildup'] = df.loc[condition_R_VOL , 'OI']

        condition_GLG_VOL = df['BuildUp'].str.contains('green' ) & (df['Vol_Change%'] > 0)
        df.loc[condition_GLG_VOL, 'LongBuildup_extra'] = df.loc[condition_GLG_VOL , 'OI']
        
        
        condition_RO_VOL = (df['BuildUp'].str.contains('red' , case=False)  | df['BuildUp'].str.contains('orange' , case=False)  ) &  (df['Vol_Change%'] > 0)
        df.loc[condition_RO_VOL , 'ShortBuildup_extra'] = df.loc[condition_RO_VOL , 'OI']

        df.fillna(value='0', inplace=True)

        #print(df)
        #df.to_csv("COALINDIA.csv") 

        return df
    
    
    def writeFuturesAnalysisData(self , symbol_string):
        link  = ""
        driver = self.login()
        #symbol_string =    "SUNPHARMA, ASIANPAINT, INDUSINDBK, JUBLFOOD, IPCALAB, AXISBANK, INDIAMART, CANBK, PFC"
        
        symbol_list = [    item.strip() for item in symbol_string.split(',')   ]

        output_list = []

        for symbol in symbol_list:
            
            df = self.getDataFrame(driver,symbol)
        
            stockname = symbol
            link = "https://opstra.definedge.com/api/openinterest/futuresopeninterest/" + symbol + "&Combined%20OpenInterest"
            
            
            cum_SpotChange =  df['SpotChange'].astype(float).sum()
            cum_Vol_Change = df['Vol_Change'].astype(float).sum()
            cum_Vol_Change_percent = df['Vol_Change%'].astype(float).sum()
            cum_OI_Change = df['OI_Change'].astype(float).sum()
            cum_LongBuildup = df['LongBuildup'].astype(float).sum()
            cum_LongBuildup_extra = df['LongBuildup_extra'].astype(float).sum()
            
            cum_ShortBuildup = df['ShortBuildup'].astype(float).sum()
            cum_ShortBuildup_extra = df['ShortBuildup_extra'].astype(float).sum()
            
            cum_OI = df['OI'].astype(float).sum()
            
            cum_BU_Change = cum_LongBuildup - cum_ShortBuildup
            cum_BU_Change_percent = (cum_LongBuildup + cum_ShortBuildup)/cum_OI * 100
            
            cum_BU_Change_extra = cum_LongBuildup_extra - cum_ShortBuildup_extra
            cum_BU_Change_percent_extra = (cum_LongBuildup_extra + cum_ShortBuildup_extra)/cum_OI * 100
            
            
            excel_filename =  self.getFilePath()+"Output/futuresOI/" +  stockname + "_futuresOI.xlsx"
            df.to_excel(excel_filename)
            wb = openpyxl.load_workbook(excel_filename)
            ws = wb['Sheet1']
            ws.insert_rows(1, amount=5)
            
            ws.cell(2,1).value = 'STOCKNAME'
            ws.cell(2,2).value = stockname
            ws.cell(2,4 ).value = link
            
            ws.cell(4,1).value = 'CumBUChange'
            ws.cell(4,2).value = cum_BU_Change
            
            ws.cell(3,1).value = 'RESULT'
            result = ''
            if cum_BU_Change < 0 :
                result  = 'BEARISH'
            else:
                result  = 'BULLISH'
            ws.cell(3,2).value = result   
            
            ws.cell(3,4).value = 'RESULT_extra'
            result_extra = ''
            if cum_BU_Change_extra < 0 :
                result_extra  = 'BEARISH'
            else:
                result_extra  = 'BULLISH'
            ws.cell(3,5).value = result_extra 
            
            ws.cell(5,1).value = 'CumBUChange%'
            ws.cell(5,2).value = cum_BU_Change_percent
            
            ws.cell(14,5).value = cum_SpotChange
            
            
            ws.cell(5, 15).value = cum_Vol_Change
            ws.cell(5,16).value = cum_Vol_Change_percent
            ws.cell(5,17).value = cum_OI_Change
            ws.cell(5,18).value = cum_LongBuildup
            ws.cell(5,19).value = cum_ShortBuildup
            ws.cell(5,20).value = cum_LongBuildup_extra
            ws.cell(5,21).value = cum_ShortBuildup_extra
            ws.cell(5,7).value = cum_OI
            
            
            wb.save(excel_filename)
            
            output_list.append([stockname, datetime.now().strftime("%Y-%m-%d_%H:%M:%S") , result,  cum_BU_Change, cum_BU_Change_percent, result_extra, cum_BU_Change_extra, cum_BU_Change_percent_extra])
            
                
        # Close the WebDriver
        driver.quit()



        output_df = pd.DataFrame(output_list)
        output_df.columns = ['STOCKNAME', 'DATE', 'RESULT', 'CumBUChange', 'CumBUChange%', 'RESULT_extra', 'CumBUChange_extra', 'CumBUChange%_extra']
        output_df.to_excel(self.getFilePath()+"Output/Opstra_Data_Screener_Analysis.xlsx")   
        
        
        # write a complete 20 day analysis report of all files 
        #self.writeCompleteFuturesAnalysisReport()
        
    def writeCompleteFuturesAnalysisReport(self):
        # Get the list of all files
        os_path = self.getFilePath()+"Output/futuresOI/"
        file_list = os.listdir(os_path)
        
        # Create an empty DataFrame to store the analysis data
        analysis_df = pd.DataFrame()

        # Loop through each file
        for file_path in file_list:
            
            if file_path.endswith('_futuresOI.xlsx') == False:
                continue    # execute loop only for excel files
            # Read the file data into a DataFrame
            wb = openpyxl.load_workbook(os_path + file_path)
            ws = wb['Sheet1']
            dict = {
                    'STOCKNAME' : ws.cell(2,2).value,
                    'DATE' : datetime.now().strftime("%Y-%m-%d_%H:%M:%S") ,
                    'RESULT' : ws.cell(3,2).value,
                    'CumBUChange' : ws.cell(4,2).value,
                    'CumBUChange%' : ws.cell(5,2).value,
                    'RESULT_extra' : '',
                    'CumBUChange_extra' : '',
                    'CumBUChange%_extra' :''               
                    
                    
                    }
            
            
            # Append the file data to the analysis DataFrame
            tempt_df = pd.DataFrame(dict, index=[0])
            analysis_df = pd.concat([analysis_df, tempt_df], ignore_index=True )
            #analysis_df = analysis_df.append(dict, ignore_index=True)
            wb.close()
            

        # Save the analysis DataFrame to a new Excel file
        analysis_df.to_excel(self.getFilePath() + "Output/Complete_Futures_Analysis_Report.xlsx")

            
    def getNextExpiryDate(self):
        month_names = ["", "JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]
        
        # Get the current date and time
        current_datetime = datetime.now()

        # Extract the current month and year
        current_month = current_datetime.month
        current_month_str = month_names[current_month]
        
        current_year_str =  str( current_datetime.year)

        expiry = current_month_str + current_year_str
        expiry = difflib.get_close_matches(expiry,self.expiry_dates)[0]
        
        return expiry
        
    
    def getTimestampList(self, driver, noOfDays = 20):
        link = "https://opstra.definedge.com/api/openinterest/futuresopeninterest/NIFTY&Combined%20OpenInterest"
        
        
        data_dict = self.get_symbol_data(driver , symbol = "NIFTY")
        df = pd.DataFrame(data_dict)
        
        timestamps_last_20_days =df["Timestamp"].tolist()
        
        for  i in range ( len(timestamps_last_20_days)  ):
            utcdatetime =    datetime.utcfromtimestamp(timestamps_last_20_days[i]/1000)
            utcdatetime = utcdatetime.replace(hour=10, minute=0, second=0, microsecond=0)
            timestamps_last_20_days[i] = int(utcdatetime.timestamp())
            
        
        return timestamps_last_20_days
        
    
    
    def writeOptionOIData(self , symbol_string):
        symbol_list = [    item.strip() for item in symbol_string.split(',')   ]

        driver = self.login()
        for symbol in symbol_list:
            self.scrapeOptionsOI(driver , symbol)
    
        driver.quit()
        
        
    def scrapeOptionsOI(self, driver ,symbol = "JSWSTEEL"):
        
        filename = self.getFilePath() + "Output/optionsOI/" + symbol + ".xlsx"
        
        
        #get the next monthly expiry date 
        expiry = self.getNextExpiryDate()
        
        # get list of timestamps for the last 30 days at 10:00 AM UTC
        timestamps_last_20_days = self.getTimestampList(driver)
        list_of_dates_from_file = []
        
        # Read the excel file and get latest scape date
        latest_scape_date = 0
        file_df = pd.DataFrame()
        if (os.path.exists(filename)):
            file_df = pd.read_excel(filename)
            list_of_dates_from_file = file_df["Timestamp"].unique().tolist()
            #latest_scape_date = file_df["Timestamp"].max()
            #latest_scape_date = int(latest_scape_date)
            #pd.close(filename)
            
            
        

        result_df = pd.DataFrame()
        #print("Next Expiry : " , expiry)
        #print("\nTimestamp for the last 30 days at 10:00 AM UTC:", timestamps_last_20_days)
        
        for time_stamp in timestamps_last_20_days:
            
            # if data for a particular date is present , copy to result_df
            if (time_stamp in list_of_dates_from_file  ):
                filtered_rows = file_df[file_df['Timestamp'] == time_stamp]
                result_df = pd.concat([result_df,filtered_rows])
                
            # else get data for that particular date
            else : 
                link = "https://opstra.definedge.com/api/openinterest/oislider/" + str(time_stamp) + "&" + symbol + "&"  + expiry
                data_dict = self.getOptionsOI(driver,link)
                df = pd.DataFrame(data_dict)
                result_df = pd.concat([result_df,df])
                #print(df)
                
        # overwrite the file or create a new one if not present
        result_df.to_excel(filename)
        
        
    
    def getOptionsOI(self,driver , link):
        df = pd.DataFrame()
        time_stamp = link.split("/")[-1] # get timestamp from the link 
        time_stamp = time_stamp.split("&")[0]
        
        
        # Navigate to the desired URL
        driver.get(link)  # json data

        # Wait for some time to allow the page to load 
        time.sleep(5)

        page_source = driver.page_source
        # Parse the HTML string with BeautifulSoup
        soup = BeautifulSoup(page_source, 'html.parser')

        # Find the <div> element with id "json-data"
        json_body = soup.find('body')
        json_div = json_body.find('div', {'hidden': 'true'})


        # Check if the <div> element exists and has content
        if json_div and json_div.string:
            # Extract the JSON data
            json_data = json_div.string

            # Parse the JSON data into a Python dictionary
            json_optionsOI_object = json.loads(json_data)["data"]
            json_futuresprice_object = json.loads(json_data)["futuresprice"]
            json_hrmaxpain_object = json.loads(json_data)["hrmaxpain"]
            json_lotsize_object = json.loads(json_data)["lotsize"]
            json_maxoic_object = json.loads(json_data)["maxoic"]
            json_maxpain_object = json.loads(json_data)["maxpain"]
            json_maxpainpos_object = json.loads(json_data)["maxpainpos"]
            json_spotprice_object = json.loads(json_data)["spotprice"]
            json_spotstrike_object = json.loads(json_data)["spotstrike"]
            json_spotstrikepos_object = json.loads(json_data)["spotstrikepos"]
            json_totalpcr_object = json.loads(json_data)["totalpcr"]
            
            df = pd.DataFrame(json_optionsOI_object)
            df["futuresprice"] = float(json_futuresprice_object)
            df["hrmaxpain"] = float(json_hrmaxpain_object)
            df["lotsize"] = float(json_lotsize_object)
            df["maxoic"] = float(json_maxoic_object)
            df["maxpain"] = float(json_maxpain_object)
            df["maxpainpos"] = int(json_maxpainpos_object)
            df["spotprice"] = float(json_spotprice_object)
            df["spotstrike"] = float(json_spotstrike_object)
            df["spotstrikepos"] = int(json_spotstrikepos_object)
            df["totalpcr"] = float(json_totalpcr_object)
            #df["Timestamp"] = int(time_stamp)
            df.insert(0, 'Timestamp', int(time_stamp))
            
            # Find the index of the row with 'StrikeValues' equal to spot strike
            index_of_spotstrike = df["spotstrikepos"][0]

            # Subset the dataframe with 15 recs before and 15 after the spot strike
            if(len(df) > 30):
                start_index = max(0, index_of_spotstrike - 15)
                end_index = min(index_of_spotstrike + 15, len(df) - 1)
                df = df.loc[start_index:end_index]
                

            # Now, 'json_object' contains the extracted JSON data
            #print(json_object)
        else:
            print("JSON data not found.")


        return df
        

        

if __name__ == '__main__':
    scraper = OpstraScraper()
    #scraper.writeOptionsData()
    #scraper.writeFuturesData()
    #scraper.writeFuturesAnalysisData()
    scraper.writeCompleteFuturesAnalysisReport()
    
